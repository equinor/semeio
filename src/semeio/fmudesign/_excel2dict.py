"""Module for reading excel file with input for generation of
a design matrix and converting to an OrderedDict that can be
read by semeio.fmudesign.DesignMatrix.generate
"""

from collections import Counter, OrderedDict
from collections.abc import Hashable, Mapping, Sequence
from pathlib import Path
from typing import Any, cast

import numpy as np
import openpyxl
import pandas as pd
import yaml


def excel2dict_design(
    input_filename: str,
    *,
    gen_input_sheet: str = "general_input",
    design_input_sheet: str = "designinput",
    default_val_sheet: str = "defaultvalues",
) -> OrderedDict[str, Any]:
    """Read excel file with input to design setup
    Currently only specification of
    onebyone design is implemented

    Args:
        input_filename (str): Name of excel input file
        gen_input_sheet (str): Sheet name for general input
        design_input_sheet (str): Sheet name for design input
        default_val_sheet (str): Sheet name for default input

    Returns:
        OrderedDict on format for DesignMatrix.generate
    """

    # Find sheets
    xlsx = openpyxl.load_workbook(input_filename, read_only=True, keep_links=False)
    gen_input_sheet = find_sheet(gen_input_sheet, names=xlsx.sheetnames)
    design_input_sheet = find_sheet(design_input_sheet, names=xlsx.sheetnames)
    default_val_sheet = find_sheet(default_val_sheet, names=xlsx.sheetnames)

    generalinput = (
        pd.read_excel(
            input_filename, gen_input_sheet, header=None, index_col=0, engine="openpyxl"
        )
        .dropna(axis=0, how="all")
        .dropna(axis=1, how="all")
    )

    design_type = str(generalinput.loc["designtype"].iloc[0])
    if design_type != "onebyone":
        raise ValueError(
            f"Generation of DesignMatrix only implemented for type 'onebyone', not {design_type}"
        )

    if "seeds" in generalinput.index:
        raise ValueError(
            "The 'seeds' parameter has been deprecated and is no longer supported. "
            "Use 'rms_seeds' instead"
        )

    return _excel2dict_onebyone(
        input_filename=input_filename,
        gen_input_sheet=gen_input_sheet,
        design_input_sheet=design_input_sheet,
        default_val_sheet=default_val_sheet,
    )


def inputdict_to_yaml(inputdict: Mapping[str, Any], filename: str) -> None:
    """Write inputdict to yaml format

    Args:
        inputdict (OrderedDict)
        filename (str): path for where to write file
    """
    with open(filename, "w", encoding="utf-8") as stream:
        yaml.dump(inputdict, stream)


def find_sheet(name: str, names: list[str]) -> str:
    """Search for Excel sheets with a soft matching. Raises ValueError if zero
    or more than one match is found.

    Examples:
    >>> find_sheet('general_input', ['generalinput', 'designinput', 'defaultinput'])
    'generalinput'
    >>> find_sheet('variable_input', ['generalinput', 'designinput', 'defaultinput'])
    Traceback (most recent call last):
      ...
    ValueError: No match for variable_input: ['generalinput', 'designinput', 'defaultinput']
    """

    def sanitize(inputstring: str) -> str:
        return inputstring.lower().strip().replace("_", "")

    found = [name_i for name_i in names if sanitize(name) == sanitize(name_i)]
    if len(found) > 1:
        raise ValueError(f"More than one match for {name}: {found}")
    elif len(found) == 0:
        raise ValueError(f"No match for {name}: {names}")
    else:
        return found[0]


def _check_designinput(dsgn_input: pd.DataFrame) -> None:
    """Checks for valid input in designinput sheet"""
    # Filter out rows where sensname has no value
    valid_sensnames = dsgn_input["sensname"].dropna()
    duplicated_mask = valid_sensnames.duplicated()

    if duplicated_mask.any():
        # Find the first duplicate to include in error message
        duplicate_name = valid_sensnames[duplicated_mask].iloc[0]
        raise ValueError(
            f"sensname '{duplicate_name}' was found on more than one row in designinput "
            "sheet. Two sensitivities can not share the same sensname. "
            "Please correct this and rerun"
        )

    # Check for duplicate parameter names within each sensname
    for sensname, df_sensname in dsgn_input.ffill().groupby("sensname"):
        param_names = list(df_sensname["param_name"])
        try:
            _raise_if_duplicates(param_names)
        except ValueError as e:
            raise ValueError(f"Duplicate param names in {sensname}\n{e}") from e


def _check_for_mixed_sensitivities(sens_name: str, sens_group: pd.DataFrame) -> None:
    """Checks for valid input in designinput sheet. A sensitivity cannot contain
    two different sensitivity types"""

    unique_types = sens_group["type"].dropna().unique()
    if len(unique_types) > 1:
        raise ValueError(
            f"The sensitivity with sensname '{sens_name}' in designinput sheet contains more "
            "than one sensitivity type. For each sensname all parameters must be "
            "specified using the same type (seed, scenario, dist, ref, background, "
            "extern)"
        )


def resolve_path(input_filename: str, reference: str) -> str:
    """The path `input_filename` is an Excel sheet, and `reference` is a cell
    value that *might* be a reference to another file. Resolve the path to
    `reference` and return.
    """
    assert str(input_filename).endswith(("xlsx", "csv"))

    # It's not a reference to another file
    if not reference.endswith(("xlsx", "csv")):
        return reference

    reference_path = Path(reference)

    # If the reference is e.g. 'C:/Users/USER/files/doe1.xlsx' => full path
    if reference_path.is_absolute():
        return str(reference_path.resolve())

    # If the reference is 'doe1.xlsx', then assume the file is in same dir
    return str(Path(input_filename).parent / reference_path)


def _excel2dict_onebyone(
    input_filename: str | Path,
    *,
    gen_input_sheet: str,
    design_input_sheet: str,
    default_val_sheet: str,
) -> OrderedDict[str, Any]:
    """Reads specification for onebyone design

    Args:
        input_filename(str or path): path to excel workbook
        gen_input_sheet (str): name of general input sheet
        design_input_sheet (str): name of design input sheet
        default_val_sheet (str): name of defaul value sheet
        sheetnames (dict): Dictionary of worksheet names to load
            information from. Supported keys: general_input, defaultvalues,
            and designinput.

    Returns:
        OrderedDict on format for DesignMatrix.generate
    """
    input_filename = str(input_filename)
    seedname = "RMS_SEED"
    inputdict: OrderedDict[str, Any] = OrderedDict()

    generalinput = pd.read_excel(
        input_filename, gen_input_sheet, header=None, index_col=0, engine="openpyxl"
    )
    generalinput.dropna(axis=0, how="all", inplace=True)
    generalinput.dropna(axis=1, how="all", inplace=True)

    inputdict["designtype"] = generalinput[1]["designtype"]

    if "rms_seeds" in generalinput.index:
        rms_seeds = str(generalinput.loc["rms_seeds"].iloc[0])
        if rms_seeds == "None":
            inputdict["seeds"] = None
        else:
            inputdict["seeds"] = resolve_path(input_filename, rms_seeds)
    else:
        inputdict["seeds"] = None

    if "repeats" not in generalinput.index:
        raise LookupError('"repeats" must be specified in general_input sheet')

    inputdict["repeats"] = generalinput.loc["repeats"].iloc[0]

    if "distribution_seed" in generalinput.index:
        distribution_seed = str(generalinput.loc["distribution_seed"].iloc[0])
        if distribution_seed == "None":
            inputdict["distribution_seed"] = None
        else:
            inputdict["distribution_seed"] = int(distribution_seed)
    else:
        inputdict["distribution_seed"] = None

    if "background" in generalinput.index:
        background = str(generalinput.loc["background"].iloc[0])
        inputdict["background"] = OrderedDict()
        if background.endswith(("csv", "xlsx")):
            inputdict["background"]["extern"] = resolve_path(input_filename, background)
        elif background == "None":
            inputdict["background"] = None
        else:
            inputdict["background"] = _read_background(input_filename, background)
    else:
        inputdict["background"] = None

    inputdict["defaultvalues"] = _read_defaultvalues(input_filename, default_val_sheet)

    inputdict["sensitivities"] = OrderedDict()
    designinput = pd.read_excel(input_filename, design_input_sheet, engine="openpyxl")
    designinput.dropna(axis=0, how="all", inplace=True)
    designinput = designinput.loc[
        :, ~designinput.columns.astype(str).str.contains("^Unnamed")
    ]

    # First column with parameter names should have spaces stripped,
    # but we need to preserve NaNs:
    not_nan_sensnames = ~designinput["sensname"].isnull()
    designinput.loc[not_nan_sensnames, "sensname"] = (
        designinput.loc[not_nan_sensnames, "sensname"].astype(str).str.strip()
    )

    _check_designinput(designinput)

    designinput["sensname"] = designinput["sensname"].ffill()

    if "dependencies" in designinput:
        valid_deps = designinput[designinput["dependencies"].notna()]
        inputdict["dependencies"] = OrderedDict()
        for row in valid_deps.itertuples():
            inputdict["dependencies"][row.param_name] = _read_dependencies(
                input_filename,
                str(row.dependencies),
                str(row.param_name),
            )

    if "decimals" in designinput:
        # Convert to numeric, then filter for integers
        numeric_decimals = pd.to_numeric(designinput["decimals"], errors="coerce")
        mask = numeric_decimals.notna() & (numeric_decimals % 1 == 0)

        valid_decimals = designinput[mask]
        inputdict["decimals"] = OrderedDict(
            {
                row.param_name: int(cast(float, row.decimals))
                for row in valid_decimals.itertuples()
            }
        )

    grouped = designinput.groupby("sensname", sort=False)

    # Read each sensitivity
    for sensname, group in grouped:
        _check_for_mixed_sensitivities(
            str(sensname),
            group,
        )

        sensdict: OrderedDict[str, Any] = OrderedDict()

        sens_type = group["type"].iloc[0]
        if sens_type in {"ref", "background"}:
            sensdict["senstype"] = sens_type

        elif sens_type == "seed":
            sensdict["seedname"] = seedname
            sensdict["senstype"] = sens_type
            if _has_value(group["param_name"].iloc[0]):
                sensdict["parameters"] = _read_constants(group)
            else:
                sensdict["parameters"] = None

        elif sens_type == "scenario":
            sensdict = _read_scenario_sensitivity(group)
            sensdict["senstype"] = sens_type

        elif sens_type == "dist":
            sensdict["senstype"] = sens_type
            sensdict["parameters"] = _read_dist_sensitivity(group)
            sensdict["correlations"] = None
            if "corr_sheet" in group:
                sensdict["correlations"] = _read_correlations(group, input_filename)

        elif sens_type == "extern":
            sensdict["extern_file"] = resolve_path(
                input_filename, str(group["extern_file"].iloc[0])
            )
            sensdict["senstype"] = sens_type
            sensdict["parameters"] = list(group["param_name"])

        else:
            raise ValueError(
                f"Sensitivity {sensname} does not have a valid sensitivity type"
            )

        if "numreal" in group and _has_value(group["numreal"].iloc[0]):
            # Using default number of realisations:
            # 'repeats' from general_input sheet
            sensdict["numreal"] = int(group["numreal"].iloc[0])

        inputdict["sensitivities"][str(sensname)] = sensdict

    return inputdict


def _read_defaultvalues(filename: str, sheetname: str) -> OrderedDict[str, Any]:
    """Reads defaultvalues, also used as values for
    reference/base case

    Args:
        filename(path): path to excel file
        sheetname (string): name of defaultsheet

    Returns:
        OrderedDict with defaultvalues (parameter, value)
    """
    default_df = pd.read_excel(
        filename, sheetname, header=0, index_col=0, engine="openpyxl"
    )

    default_df.dropna(axis=0, how="all", inplace=True)
    default_df = default_df.loc[
        :, ~default_df.columns.astype(str).str.contains("^Unnamed")
    ]

    if default_df.empty:
        return OrderedDict()

    # Strip leading/trailing spaces from parameter names such that
    # for example "  PARAM" and "PARAM" are treated as duplicates.
    default_df.index = default_df.index.str.strip()

    # Check for duplicates and raise error if found
    duplicates = default_df.index.duplicated(keep=False)
    if duplicates.any():
        duplicate_names = default_df.index[duplicates].unique()
        raise ValueError(
            f"Duplicate parameter names found in sheet '{sheetname}': "
            f"{', '.join(duplicate_names)}. All parameter names must be unique."
        )
    return OrderedDict(default_df.iloc[:, 0].to_dict())


def _read_dependencies(
    filename: str, sheetname: str, from_parameter: str
) -> OrderedDict[str, Any]:
    """Reads parameters that are set from other parameters

    Args:
        filename(path): path to excel file
        sheetname (string): name of dependency sheet

    Returns:
        OrderedDict with design parameter, dependent parameters
        and values
    """
    depend_dict: OrderedDict[str, Any] = OrderedDict()
    depend_df = pd.read_excel(
        filename, sheetname, dtype=str, na_values="", engine="openpyxl"
    )
    depend_df.dropna(axis=0, how="all", inplace=True)
    depend_df = depend_df.loc[
        :, ~depend_df.columns.astype(str).str.contains("^Unnamed")
    ]

    if from_parameter in depend_df:
        depend_dict["from_values"] = depend_df[from_parameter].tolist()
        depend_dict["to_params"] = OrderedDict()
        for key in depend_df:
            if key != from_parameter:
                depend_dict["to_params"][key] = depend_df[key].tolist()
    else:
        raise ValueError(
            f"Parameter {from_parameter} specified to have derived parameters, "
            f"but the sheet specifying the dependencies {sheetname} does "
            "not contain the input parameter. "
        )
    return depend_dict


def _read_background(inp_filename: str, bck_sheet: str) -> OrderedDict[str, Any]:
    """Reads excel sheet with background parameters and distributions

    Args:
        inp_filename (path): path to Excel workbook
        bck_sheet (str): name of sheet with background parameters

    Returns:
        OrderedDict with parameter names and distributions
    """
    backdict: OrderedDict[str, Any] = OrderedDict()
    paramdict: OrderedDict[str, Any] = OrderedDict()
    bck_input = pd.read_excel(inp_filename, bck_sheet, engine="openpyxl")
    bck_input.dropna(axis=0, how="all", inplace=True)
    bck_input = bck_input.loc[
        :, ~bck_input.columns.astype(str).str.contains("^Unnamed")
    ]

    backdict["correlations"] = None
    if "corr_sheet" in bck_input:
        backdict["correlations"] = _read_correlations(bck_input, inp_filename)

    if "dist_param1" not in bck_input.columns.values:
        bck_input["dist_param1"] = float("NaN")
    if "dist_param2" not in bck_input.columns.values:
        bck_input["dist_param2"] = float("NaN")
    if "dist_param3" not in bck_input.columns.values:
        bck_input["dist_param3"] = float("NaN")
    if "dist_param4" not in bck_input.columns.values:
        bck_input["dist_param4"] = float("NaN")

    for row in bck_input.itertuples():
        if not _has_value(row.param_name):
            raise ValueError(
                "Background parameters specified "
                "where one line has empty parameter "
                "name "
            )
        if not _has_value(row.dist_param1):
            raise ValueError(
                f"Parameter {row.param_name} has been input "
                "in background sheet but with empty "
                "first distribution parameter "
            )
        if not _has_value(row.dist_param2) and _has_value(row.dist_param3):
            raise ValueError(
                f"Parameter {row.param_name} has been input in "
                "background sheet with "
                'value for "dist_param3" while '
                '"dist_param2" is empty. This is not '
                "allowed"
            )
        if not _has_value(row.dist_param3) and _has_value(row.dist_param4):
            raise ValueError(
                f"Parameter {row.param_name} has been input in "
                "background sheet with "
                'value for "dist_param4" while '
                '"dist_param3" is empty. This is not '
                "allowed"
            )
        distparams = [
            item
            for item in [
                row.dist_param1,
                row.dist_param2,
                row.dist_param3,
                row.dist_param4,
            ]
            if _has_value(item)
        ]
        if "corr_sheet" in bck_input:
            corrsheet = None if not _has_value(row.corr_sheet) else row.corr_sheet
        else:
            corrsheet = None
        paramdict[str(row.param_name)] = [str(row.dist_name), distparams, corrsheet]
    backdict["parameters"] = paramdict

    if "decimals" in bck_input:
        decimals: OrderedDict[str, Any] = OrderedDict()
        for row in bck_input.itertuples():
            if _has_value(row.decimals) and _is_int(row.decimals):  # type: ignore[arg-type]
                decimals[row.param_name] = int(row.decimals)  # type: ignore[arg-type, index]
        backdict["decimals"] = decimals

    return backdict


def _read_scenario_sensitivity(sensgroup: pd.DataFrame) -> OrderedDict[str, Any]:
    """Reads parameters and values
    for scenario sensitivities
    """
    sdict: OrderedDict[str, Any] = OrderedDict()
    sdict["cases"] = OrderedDict()
    casedict1: OrderedDict[str, Any] = OrderedDict()
    casedict2: OrderedDict[str, Any] = OrderedDict()

    if not _has_value(sensgroup["senscase1"].iloc[0]):
        raise ValueError(
            "Sensitivity {} has been input "
            "as a scenario sensitivity, but "
            "without a name in senscase1 column.".format(sensgroup["sensname"].iloc[0])
        )

    for row in sensgroup.itertuples():
        if not _has_value(row.param_name):
            raise ValueError(
                f"Scenario sensitivity {row.sensname} specified "
                "where one line has empty parameter "
                "name "
            )
        if not _has_value(row.value1):
            raise ValueError(
                f"Parameter {row.param_name} har been input "
                'as type "scenario" but with empty '
                "value in value1 column "
            )
        casedict1[str(row.param_name)] = row.value1

    if _has_value(sensgroup["senscase2"].iloc[0]):
        for row in sensgroup.itertuples():
            if not _has_value(row.value2):
                raise ValueError(
                    "Sensitivity {} has been input "
                    "with a name in senscase2 column "
                    "but without a value for parameter {} "
                    "in value2 column.".format(
                        sensgroup["sensname"].iloc[0], row.param_name
                    )
                )
            casedict2[str(row.param_name)] = row.value2
        sdict["cases"][str(sensgroup["senscase1"].iloc[0])] = casedict1
        sdict["cases"][str(sensgroup["senscase2"].iloc[0])] = casedict2
    else:
        for row in sensgroup.itertuples():
            if _has_value(row.value2):
                raise ValueError(
                    "Sensitivity {} has been input "
                    "with a value for parameter {} "
                    "in value2 column "
                    "but without a name for the scenario "
                    "in senscase2 column.".format(
                        sensgroup["sensname"].iloc[0], row.param_name
                    )
                )
        sdict["cases"][str(sensgroup["senscase1"].iloc[0])] = casedict1
    return sdict


def _read_constants(sensgroup: pd.DataFrame) -> OrderedDict[str, Any]:
    """Reads constants to be used together with
    seed sensitivity"""
    if "dist_param1" not in sensgroup.columns.values:
        sensgroup["dist_param1"] = float("NaN")
    paramdict: OrderedDict[str, Any] = OrderedDict()
    for row in sensgroup.itertuples():
        if not _has_value(row.dist_param1):
            raise ValueError(
                f"Parameter name {row.param_name} has been input "
                'in a sensitivity of type "seed". \n'
                f"If {row.param_name} was meant to be the name of "
                "the seed parameter, this is "
                "unfortunately not allowed. "
                "The seed parameter name is standardised "
                "to RMS_SEED and should not be specified.\n "
                "If you instead meant to specify a constant "
                "value for another parameter in the seed "
                'sensitivity, please remember "const" in '
                'dist_name and a value in "dist_param1". '
            )
        distparams = row.dist_param1
        paramdict[str(row.param_name)] = [str(row.dist_name), distparams]
    return paramdict


def _read_dist_sensitivity(sensgroup: pd.DataFrame) -> OrderedDict[str, Any]:
    """Reads parameters and distributions
    for monte carlo sensitivities
    """
    if "dist_param1" not in sensgroup.columns.values:
        sensgroup["dist_param1"] = float("NaN")
    if "dist_param2" not in sensgroup.columns.values:
        sensgroup["dist_param2"] = float("NaN")
    if "dist_param3" not in sensgroup.columns.values:
        sensgroup["dist_param3"] = float("NaN")
    if "dist_param4" not in sensgroup.columns.values:
        sensgroup["dist_param4"] = float("NaN")
    paramdict: OrderedDict[str, Any] = OrderedDict()
    for row in sensgroup.itertuples():
        if not _has_value(row.param_name):
            raise ValueError(
                f"Dist sensitivity {row.sensname} specified "
                "where one line has empty parameter "
                "name "
            )
        if not _has_value(row.dist_param1):
            raise ValueError(
                f"Parameter {row.param_name} has been input "
                'as type "dist" but with empty '
                "first distribution parameter "
            )
        if not _has_value(row.dist_param2) and _has_value(row.dist_param3):
            raise ValueError(
                f"Parameter {row.param_name} has been input with "
                'value for "dist_param3" while '
                '"dist_param2" is empty. This is not '
                "allowed"
            )
        if not _has_value(row.dist_param3) and _has_value(row.dist_param4):
            raise ValueError(
                f"Parameter {row.param_name} has been input with "
                'value for "dist_param4" while '
                '"dist_param3" is empty. This is not '
                "allowed"
            )
        distparams = [
            item
            for item in [
                row.dist_param1,
                row.dist_param2,
                row.dist_param3,
                row.dist_param4,
            ]
            if _has_value(item)
        ]
        if "corr_sheet" in sensgroup:
            corrsheet = None if not _has_value(row.corr_sheet) else row.corr_sheet
        else:
            corrsheet = None
        paramdict[str(row.param_name)] = [str(row.dist_name), distparams, corrsheet]

    return paramdict


def _read_correlations(
    sensgroup: pd.DataFrame, inputfile: str
) -> OrderedDict[str, Any] | None:
    if "corr_sheet" in sensgroup:
        if not sensgroup["corr_sheet"].dropna().empty:
            correlations: OrderedDict[str, Any] = OrderedDict()
            correlations["inputfile"] = inputfile
            correlations["sheetnames"] = []
            for _index, row in sensgroup.iterrows():
                if (
                    _has_value(row["corr_sheet"])
                    and row["corr_sheet"] not in correlations["sheetnames"]
                ):
                    correlations["sheetnames"].append(row["corr_sheet"])
        else:
            return None
    else:
        return None

    return correlations


def _has_value(value: Any) -> bool:  # noqa: ANN401
    """Returns False only if the argument is np.nan"""
    try:
        return not np.isnan(value)
    except TypeError:
        return True


def _is_int(teststring: str) -> bool:
    """Test if a string can be parsed as a float"""
    try:
        if not np.isnan(int(teststring)):
            return (float(teststring) % 1) == 0
        return False  # It was a "number", but it was NaN.
    except ValueError:
        return False


def _raise_if_duplicates(container: Sequence[Hashable]) -> None:
    """Raises a descriptive error if there are duplicates in the container."""
    duplicates = {k: v for (k, v) in Counter(container).items() if v > 1}
    if duplicates:
        raise ValueError(f"Duplicates with counts: {duplicates}")
