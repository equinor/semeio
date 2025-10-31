"""This module contains functions for reading Excel config files.
These are converted to a dict-of-dicts representation, then they are used
by the DesignMatrix class to generate design matrices.
"""

import collections
import contextlib
from collections import Counter
from collections.abc import Hashable, Sequence
from pathlib import Path
from typing import Any, cast

import numpy as np
import openpyxl
import pandas as pd
import yaml

from semeio.fmudesign.design_distributions import read_correlations
from semeio.fmudesign.utils import seeds_from_extern


def excel_to_dict(
    input_filename: str,
    *,
    gen_input_sheet: str = "general_input",
    design_input_sheet: str = "designinput",
    default_val_sheet: str = "defaultvalues",
) -> dict[str, Any]:
    """Read excel file with input to design setup

    Args:
        input_filename (str): Name of excel input file
        gen_input_sheet (str): Sheet name for general input
        design_input_sheet (str): Sheet name for design input
        default_val_sheet (str): Sheet name for default input

    Returns:
        dict on format for DesignMatrix.generate
    """
    # To be backwards compatible, we do not change the input arg names
    general_input_sheet = gen_input_sheet
    default_values_sheet = default_val_sheet

    # Find sheets
    _assert_no_merged_cells(input_filename)
    xlsx = openpyxl.load_workbook(input_filename, read_only=True, keep_links=False)
    general_input_sheet = find_sheet(general_input_sheet, names=xlsx.sheetnames)
    design_input_sheet = find_sheet(design_input_sheet, names=xlsx.sheetnames)
    default_values_sheet = find_sheet(default_values_sheet, names=xlsx.sheetnames)

    generalinput = (
        pd.read_excel(
            input_filename,
            general_input_sheet,
            header=None,
            index_col=0,
            engine="openpyxl",
        )
        .dropna(axis=0, how="all")
        .dropna(axis=1, how="all")
        .loc[:, 1]
        .to_dict()
    )

    if (design_type := generalinput.get("designtype")) != "onebyone":
        raise ValueError(
            f"Generation of DesignMatrix only implemented for type 'onebyone', not {design_type}"
        )

    return _excel_to_dict_onebyone(
        input_filename=input_filename,
        general_input_sheet=general_input_sheet,
        design_input_sheet=design_input_sheet,
        default_values_sheet=default_values_sheet,
    )


def inputdict_to_yaml(inputdict: dict[str, Any], filename: str) -> None:
    """Write inputdict to yaml format

    Args:
        inputdict (dict)
        filename (str): name of output file
    """
    with open(filename, "w", encoding="utf-8") as stream:
        yaml.dump(inputdict, stream)


def find_sheet(name: str, names: list[str]) -> str:
    """Search for Excel sheets with a soft matching. Raises ValueError if zero
    or more than one match is found.

    Examples:
    >>> find_sheet('general_input', ['generalinput', 'designinput', 'defaultinput'])
    'generalinput'
    >>> find_sheet('variable_input', ['generalinput', 'designinput', 'defaultinput'])
    Traceback (most recent call last):
      ...
    ValueError: No match for variable_input: ['generalinput', 'designinput', 'defaultinput']
    """

    def sanitize(inputstring: str) -> str:
        return inputstring.lower().strip().replace("_", "")

    found = [name_i for name_i in names if sanitize(name) == sanitize(name_i)]
    if len(found) > 1:
        raise ValueError(f"More than one match for {name}: {found}")
    elif len(found) == 0:
        raise ValueError(f"No match for {name}: {names}")
    else:
        return found[0]


def _check_designinput(dsgn_input: pd.DataFrame) -> None:
    """Checks for valid input in designinput sheet"""
    # Filter out rows where sensname has no value
    valid_sensnames = dsgn_input["sensname"].dropna()
    duplicated_mask = valid_sensnames.duplicated()

    if duplicated_mask.any():
        # Find the first duplicate to include in error message
        duplicate_name = valid_sensnames[duplicated_mask].iloc[0]
        raise ValueError(
            f"sensname '{duplicate_name}' was found on more than one row in designinput "
            "sheet. Two sensitivities can not share the same sensname. "
            "Please correct this and rerun"
        )

    # Check for duplicate parameter names within each sensname
    for sensname, df_sensname in dsgn_input.ffill().groupby("sensname"):
        param_names = list(df_sensname["param_name"])
        try:
            _raise_if_duplicates(param_names)
        except ValueError as e:
            raise ValueError(f"Duplicate param names in {sensname}\n{e}") from e


def _check_for_mixed_sensitivities(sens_name: str, sens_group: pd.DataFrame) -> None:
    """Checks for valid input in designinput sheet. A sensitivity cannot contain
    two different sensitivity types"""

    unique_types = sens_group["type"].dropna().unique()
    if len(unique_types) > 1:
        raise ValueError(
            f"The sensitivity with sensname '{sens_name}' in designinput sheet contains more "
            "than one sensitivity type. For each sensname all parameters must be "
            "specified using the same type (seed, scenario, dist, ref, background, "
            "extern)"
        )


def resolve_path(input_filename: str, reference: str | None) -> str | None:
    """The path `input_filename` is an Excel sheet, and `reference` is a cell
    value that *might* be a reference to another file. Resolve the path to
    `reference` and return. If no such file exists, return `reference`.
    """
    # The reference is None, so just return it back
    if reference is None:
        return reference

    # It's a string but not a reference to another file
    if not str(reference).endswith(("xlsx", "csv")):
        return reference

    # If the reference is e.g. 'C:/Users/USER/files/doe1.xlsx'
    reference_path = Path(reference)
    if reference_path.is_absolute() and reference_path.exists():
        return str(reference_path.resolve())

    # If the reference is e.g. 'doe1.xlsx'
    full_path = Path(input_filename).parent / reference_path
    if full_path.exists():
        return str(full_path.resolve())

    if reference_path.exists():
        return str(reference_path.resolve())

    raise ValueError(f"Failed to resolve path for file: {reference}")


def _excel_to_dict_onebyone(
    input_filename: str,
    *,
    general_input_sheet: str,
    design_input_sheet: str,
    default_values_sheet: str,
) -> dict[str, Any]:
    """Reads configuration from Excel file for a onebyone design matrix.

    Args:
        input_filename (str): Name of excel workbook
        general_input_sheet (str): name of general input sheet
        design_input_sheet (str): name of design input sheet
        default_values_sheet (str): name of default value sheet

    Returns:
        dict on format for DesignMatrix.generate
    """
    output: dict[str, Any] = {
        "input_file": input_filename
    }  # This is the config that we read and return

    # Read the general input sheet to a dictionary
    generalinput = (
        pd.read_excel(
            input_filename,
            general_input_sheet,
            header=None,
            engine="openpyxl",
        )
        .dropna(axis=0, how="all")
        .dropna(axis=1, how="all")
        .set_index(0)
        .loc[:, 1]
        .to_dict()
    )

    def parse_value(value: object) -> object:
        if pd.isna(value):  # type: ignore[call-overload]
            return None
        elif isinstance(value, str):
            return value.strip()
        return value

    # Convert NaN values to None and strip other values
    generalinput = {
        key.strip(): parse_value(value) for (key, value) in generalinput.items()
    }

    # Check that there are no wrong keys or typos, e.g. 'repets'
    ALLOWED_KEYS = {
        "designtype",
        "repeats",
        "correlation_iterations",
        "distribution_seed",
        "rms_seeds",
        "background",
    }
    extra_keys = set(generalinput.keys()) - set(ALLOWED_KEYS)
    if extra_keys:
        msg = "In the general input sheet, the following parameter(s) are not"
        msg += f"recognized and cannot be parsed:\n{extra_keys!r}\nAllowed keys:{ALLOWED_KEYS!r}"
        raise LookupError(msg)

    # Copy keys over if they exist
    keys = ["designtype", "repeats", "correlation_iterations", "distribution_seed"]
    for key in keys:
        if key not in generalinput:
            continue
        output[key] = generalinput[key]

    # Copy the 'rms_seeds' key over. It is called 'seeds' further down in
    # the code for historical reasons.
    key = "seeds"
    with contextlib.suppress(KeyError):
        output[key] = generalinput["rms_seeds"]

    # If 'seeds' / 'rms_seed' is a file, then read it
    if key in output:
        maybe_path = resolve_path(input_filename, output[key])
        if isinstance(maybe_path, str) and Path(maybe_path).exists():
            output[key] = seeds_from_extern(maybe_path)

    # If 'background' is a file, then read it
    key = "background"
    output[key] = {}
    try:
        value = str(generalinput[key])
        if value.endswith(("csv", "xlsx")):
            output[key]["extern"] = resolve_path(input_filename, value)
        else:
            output[key] = _read_background(input_filename, value)
    except KeyError:
        output[key] = None
    except ValueError:
        output[key] = generalinput[key]

    output["defaultvalues"] = _read_defaultvalues(input_filename, default_values_sheet)

    output["sensitivities"] = {}
    designinput = (
        pd.read_excel(input_filename, design_input_sheet, engine="openpyxl")
        .dropna(axis=0, how="all")
        .loc[:, lambda df: ~df.columns.astype(str).str.contains("^Unnamed")]
    )

    # Strip strings in column 'sensname' while preserving NaN values
    designinput = designinput.assign(sensname=lambda df: df["sensname"].str.strip())

    _check_designinput(designinput)

    designinput["sensname"] = designinput["sensname"].ffill()

    if "decimals" in designinput:
        # Convert to numeric, then filter for integers
        numeric_decimals = pd.to_numeric(designinput["decimals"], errors="coerce")
        mask = numeric_decimals.notna() & (numeric_decimals % 1 == 0)

        valid_decimals = designinput[mask]
        output["decimals"] = {
            row.param_name: int(cast(float, row.decimals))
            for row in valid_decimals.itertuples()
        }

    grouped = designinput.groupby("sensname", sort=False)

    # Read each sensitivity
    for sensname, group in grouped:
        _check_for_mixed_sensitivities(
            str(sensname),
            group,
        )

        sensdict: dict[str, Any] = {}

        sens_type = group["type"].iloc[0]
        if sens_type in {"ref", "background"}:
            sensdict["senstype"] = sens_type

        elif sens_type == "seed":
            sensdict["seedname"] = "RMS_SEED"
            sensdict["senstype"] = sens_type
            if _has_value(group["param_name"].iloc[0]):
                sensdict["parameters"] = _read_constants(group)
            else:
                sensdict["parameters"] = None

        elif sens_type == "scenario":
            sensdict = _read_scenario_sensitivity(group)
            sensdict["senstype"] = sens_type

        elif sens_type == "dist":
            sensdict["senstype"] = sens_type
            sensdict["parameters"] = _read_dist_sensitivity(group)
            sensdict["correlations"] = None
            if "corr_sheet" in group:
                sensdict["correlations"] = _read_correlations(group, input_filename)

        elif sens_type == "extern":
            sensdict["extern_file"] = resolve_path(
                input_filename, str(group["extern_file"].iloc[0])
            )
            sensdict["senstype"] = sens_type
            sensdict["parameters"] = list(group["param_name"])

        else:
            raise ValueError(
                f"Sensitivity {sensname} does not have a valid sensitivity type"
            )

        if "numreal" in group and _has_value(group["numreal"].iloc[0]):
            # Using default number of realisations:
            # 'repeats' from general_input sheet
            sensdict["numreal"] = int(group["numreal"].iloc[0])

        # If this sensitivity has dependencies, then get them from sheet
        sensdict["dependencies"] = {}
        if "dependencies" in group:
            # Get all dependencies in this sensitivity
            valid_deps = group[group["dependencies"].notna()]
            dependencies_dict = {}

            # For each dependency, get the mapping
            for row in valid_deps.itertuples():
                dependencies_dict[row.param_name] = _read_dependencies(
                    filename=input_filename,
                    sheetname=str(row.dependencies),
                    from_parameter=str(row.param_name),
                )
            sensdict["dependencies"] = dependencies_dict

        # Add this sensitivity to the sensitivities
        output["sensitivities"][str(sensname)] = sensdict

    return output


def _read_defaultvalues(filename: str, sheetname: str) -> dict[str, Any]:
    """Reads defaultvalues, also used as values for
    reference/base case

    Args:
        filename (str): Name of excel file
        sheetname (string): name of defaultsheet

    Returns:
        dict with defaultvalues (parameter, value)
    """
    default_df = (
        pd.read_excel(filename, sheetname, header=0, index_col=0, engine="openpyxl")
        .dropna(axis=0, how="all")
        # Drop all unnamed columns from the df
        .loc[:, lambda df: ~df.columns.astype(str).str.contains("^Unnamed")]
    )

    if default_df.empty:
        return {}

    # Strip leading/trailing spaces from parameter names such that
    # for example "  PARAM" and "PARAM" are treated as duplicates.
    default_df.index = default_df.index.str.strip()

    # Check for duplicates and raise error if found
    duplicates = default_df.index.duplicated(keep=False)
    if duplicates.any():
        duplicate_names = default_df.index[duplicates].unique()
        raise ValueError(
            f"Duplicate parameter names found in sheet '{sheetname}': "
            f"{', '.join(duplicate_names)}. All parameter names must be unique."
        )
    return default_df.iloc[:, 0].to_dict()


def _read_dependencies(
    *, filename: str, sheetname: str, from_parameter: str
) -> dict[str, Any]:
    """Reads parameters that are set from other parameters

    Args:
        filename(str): name of excel file
        sheetname (string): name of dependency sheet
        from_parameter (string): parameter name to map from

    Returns:
        dict with design parameter, dependent parameters
        and values
    """
    depend_dict: dict[str, Any] = {}
    depend_df = (
        pd.read_excel(filename, sheetname, dtype=str, na_values="", engine="openpyxl")
        .dropna(axis=0, how="all")
        .loc[:, lambda df: ~df.columns.astype(str).str.contains("^Unnamed")]
    )

    if from_parameter in depend_df:
        depend_dict["from_values"] = depend_df[from_parameter].tolist()
        depend_dict["to_params"] = {}
        for key in depend_df:
            if key != from_parameter:
                depend_dict["to_params"][key] = depend_df[key].tolist()
    else:
        raise ValueError(
            f"Parameter {from_parameter} specified to have derived parameters, "
            f"but the sheet specifying the dependencies {sheetname} does "
            "not contain the input parameter. "
        )
    return depend_dict


def _read_background(inp_filename: str, bck_sheet: str) -> dict[str, Any]:
    """Reads excel sheet with background parameters and distributions

    Args:
        inp_filename (str): name of Excel workbook
        bck_sheet (str): name of sheet with background parameters

    Returns:
        dict with parameter names and distributions
    """
    backdict: dict[str, Any] = {}
    paramdict: dict[str, Any] = {}
    bck_input = (
        pd.read_excel(inp_filename, bck_sheet, engine="openpyxl")
        .dropna(axis=0, how="all")
        .loc[:, lambda df: ~df.columns.astype(str).str.contains("^Unnamed")]
    )

    backdict["correlations"] = None
    if "corr_sheet" in bck_input:
        backdict["correlations"] = _read_correlations(bck_input, inp_filename)

    if "dist_param1" not in bck_input.columns.values:
        bck_input["dist_param1"] = float("NaN")
    if "dist_param2" not in bck_input.columns.values:
        bck_input["dist_param2"] = float("NaN")
    if "dist_param3" not in bck_input.columns.values:
        bck_input["dist_param3"] = float("NaN")
    if "dist_param4" not in bck_input.columns.values:
        bck_input["dist_param4"] = float("NaN")

    for row in bck_input.itertuples():
        if not _has_value(row.param_name):
            raise ValueError(
                "Background parameters specified "
                "where one line has empty parameter "
                "name "
            )
        if not _has_value(row.dist_param1):
            raise ValueError(
                f"Parameter {row.param_name} has been input "
                "in background sheet but with empty "
                "first distribution parameter "
            )
        if not _has_value(row.dist_param2) and _has_value(row.dist_param3):
            raise ValueError(
                f"Parameter {row.param_name} has been input in "
                "background sheet with "
                'value for "dist_param3" while '
                '"dist_param2" is empty. This is not '
                "allowed"
            )
        if not _has_value(row.dist_param3) and _has_value(row.dist_param4):
            raise ValueError(
                f"Parameter {row.param_name} has been input in "
                "background sheet with "
                'value for "dist_param4" while '
                '"dist_param3" is empty. This is not '
                "allowed"
            )
        distparams = [
            item
            for item in [
                row.dist_param1,
                row.dist_param2,
                row.dist_param3,
                row.dist_param4,
            ]
            if _has_value(item)
        ]
        if "corr_sheet" in bck_input:
            corrsheet = None if not _has_value(row.corr_sheet) else row.corr_sheet
        else:
            corrsheet = None
        paramdict[str(row.param_name)] = [str(row.dist_name), distparams, corrsheet]
    backdict["parameters"] = paramdict

    if "decimals" in bck_input:
        decimals: dict[str, Any] = {}
        for row in bck_input.itertuples():
            if _has_value(row.decimals) and _is_int(row.decimals):  # type: ignore[arg-type]
                decimals[row.param_name] = int(row.decimals)  # type: ignore[arg-type, index]
        backdict["decimals"] = decimals

    return backdict


def _read_scenario_sensitivity(sensgroup: pd.DataFrame) -> dict[str, Any]:
    """Reads parameters and values
    for scenario sensitivities
    """
    sdict: dict[str, Any] = {}
    sdict["cases"] = {}
    casedict1: dict[str, Any] = {}
    casedict2: dict[str, Any] = {}

    if not _has_value(sensgroup["senscase1"].iloc[0]):
        raise ValueError(
            "Sensitivity {} has been input "
            "as a scenario sensitivity, but "
            "without a name in senscase1 column.".format(sensgroup["sensname"].iloc[0])
        )

    for row in sensgroup.itertuples():
        if not _has_value(row.param_name):
            raise ValueError(
                f"Scenario sensitivity {row.sensname} specified "
                "where one line has empty parameter "
                "name "
            )
        if not _has_value(row.value1):
            raise ValueError(
                f"Parameter {row.param_name} has been input "
                'as type "scenario" but with empty '
                "value in value1 column "
            )
        casedict1[str(row.param_name)] = row.value1

    if _has_value(sensgroup["senscase2"].iloc[0]):
        for row in sensgroup.itertuples():
            if not _has_value(row.value2):
                raise ValueError(
                    "Sensitivity {} has been input "
                    "with a name in senscase2 column "
                    "but without a value for parameter {} "
                    "in value2 column.".format(
                        sensgroup["sensname"].iloc[0], row.param_name
                    )
                )
            casedict2[str(row.param_name)] = row.value2
        sdict["cases"][str(sensgroup["senscase1"].iloc[0])] = casedict1
        sdict["cases"][str(sensgroup["senscase2"].iloc[0])] = casedict2
    else:
        for row in sensgroup.itertuples():
            if _has_value(row.value2):
                raise ValueError(
                    "Sensitivity {} has been input "
                    "with a value for parameter {} "
                    "in value2 column "
                    "but without a name for the scenario "
                    "in senscase2 column.".format(
                        sensgroup["sensname"].iloc[0], row.param_name
                    )
                )
        sdict["cases"][str(sensgroup["senscase1"].iloc[0])] = casedict1
    return sdict


def _read_constants(sensgroup: pd.DataFrame) -> dict[str, Any]:
    """Reads constants to be used together with
    seed sensitivity"""
    if "dist_param1" not in sensgroup.columns.values:
        sensgroup["dist_param1"] = float("NaN")
    paramdict: dict[str, Any] = {}
    for row in sensgroup.itertuples():
        if not _has_value(row.dist_param1):
            raise ValueError(
                f"Parameter name {row.param_name} has been input "
                'in a sensitivity of type "seed". \n'
                f"If {row.param_name} was meant to be the name of "
                "the seed parameter, this is "
                "unfortunately not allowed. "
                "The seed parameter name is standardised "
                "to RMS_SEED and should not be specified.\n "
                "If you instead meant to specify a constant "
                "value for another parameter in the seed "
                'sensitivity, please remember "const" in '
                'dist_name and a value in "dist_param1". '
            )
        distparams = row.dist_param1
        paramdict[str(row.param_name)] = [str(row.dist_name), distparams]
    return paramdict


def _read_dist_sensitivity(sensgroup: pd.DataFrame) -> dict[str, Any]:
    """Reads parameters and distributions
    for monte carlo sensitivities
    """
    if "dist_param1" not in sensgroup.columns.values:
        sensgroup["dist_param1"] = float("NaN")
    if "dist_param2" not in sensgroup.columns.values:
        sensgroup["dist_param2"] = float("NaN")
    if "dist_param3" not in sensgroup.columns.values:
        sensgroup["dist_param3"] = float("NaN")
    if "dist_param4" not in sensgroup.columns.values:
        sensgroup["dist_param4"] = float("NaN")
    paramdict: dict[str, Any] = {}
    for row in sensgroup.itertuples():
        if not _has_value(row.param_name):
            raise ValueError(
                f"Dist sensitivity {row.sensname} specified "
                "where one line has empty parameter "
                "name "
            )
        if not _has_value(row.dist_param1):
            raise ValueError(
                f"Parameter {row.param_name} has been input "
                'as type "dist" but with empty '
                "first distribution parameter "
            )
        if not _has_value(row.dist_param2) and _has_value(row.dist_param3):
            raise ValueError(
                f"Parameter {row.param_name} has been input with "
                'value for "dist_param3" while '
                '"dist_param2" is empty. This is not '
                "allowed"
            )
        if not _has_value(row.dist_param3) and _has_value(row.dist_param4):
            raise ValueError(
                f"Parameter {row.param_name} has been input with "
                'value for "dist_param4" while '
                '"dist_param3" is empty. This is not '
                "allowed"
            )
        distparams = [
            item
            for item in [
                row.dist_param1,
                row.dist_param2,
                row.dist_param3,
                row.dist_param4,
            ]
            if _has_value(item)
        ]
        if "corr_sheet" in sensgroup:
            corrsheet = None if not _has_value(row.corr_sheet) else row.corr_sheet
        else:
            corrsheet = None
        paramdict[str(row.param_name)] = [str(row.dist_name), distparams, corrsheet]

    return paramdict


def _read_correlations(
    sensgroup: pd.DataFrame, inputfile: str
) -> dict[str, Any] | None:
    """Parse correlation information from a sensitivity group."""

    # No correlation sheet column exists
    if "corr_sheet" not in sensgroup.columns:
        return None

    # The column exists, but it is all blank
    if sensgroup["corr_sheet"].dropna().empty:
        return None

    correlations: dict[str, Any] = {"inputfile": inputfile}

    # Create a mapping 'corr_to_params' like:
    # {'corr1': ['var_A', 'var_B', ...], ...}
    corr_to_params = collections.defaultdict(list)
    for _, row in sensgroup.iterrows():
        if not _has_value(row["corr_sheet"]):
            continue
        corr_to_params[row["corr_sheet"]].append(row["param_name"])

    # Open the correlation sheet and peek at it
    # We want to verify that if variables ['A', 'B'] point to the corr sheet,
    # then exactly those variables are also defined in the sheet
    for corr_sheet, parameters in corr_to_params.items():
        df_corr = read_correlations(excel_filename=inputfile, corr_sheet=corr_sheet)
        if set(df_corr.columns) != set(parameters):
            sensname = sensgroup["sensname"].iloc[0]
            msg = f"Mismatch between parameters in sensitivity group {sensname!r} "
            msg += f"pointing to\ncorrelation sheet {corr_sheet!r} and "
            msg += "parameters specified in that correlation sheet.\n"
            msg += f"Parameters in sensitivity group: {sorted(set(parameters))}\n"
            msg += f"Parameters in correlation sheet: {sorted(set(df_corr.columns))}\n"
            msg += "These parameters must be specified one-to-one."
            raise ValueError(msg)

    correlations["sheetnames"] = list(set(corr_to_params.keys()))

    return correlations


def _has_value(value: Any) -> bool:  # noqa: ANN401
    """Returns False only if the argument is np.nan"""
    try:
        return not np.isnan(value)
    except TypeError:
        return True


def _is_int(teststring: str) -> bool:
    """Test if a string can be parsed as a float"""
    try:
        if not np.isnan(int(teststring)):
            return (float(teststring) % 1) == 0
        return False  # It was a "number", but it was NaN.
    except ValueError:
        return False


def _raise_if_duplicates(container: Sequence[Hashable]) -> None:
    """Raises a descriptive error if there are duplicates in the container."""
    duplicates = {k: v for (k, v) in Counter(container).items() if v > 1}
    if duplicates:
        raise ValueError(f"Duplicates with counts: {duplicates}")


def _assert_no_merged_cells(input_filename: str) -> None:
    """Raises an exception if any merged cells exist, else returns None."""

    workbook = openpyxl.load_workbook(input_filename)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        merged_ranges = list(worksheet.merged_cells.ranges)
        if merged_ranges:
            raise Exception(
                f"Merged cells are not allowed. Found merged cell in {input_filename} at sheet '{sheet_name}'.\n"
                f"Found {len(merged_ranges)} merged cell range(s): {merged_ranges}"
            )
