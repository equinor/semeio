import openpyxl
import pytest

from semeio.fmudesign._workbook import render_workbook

MINIMAL_SPEC = {
    "version": 1,
    "general": {
        "repeats": 1,
        "rms_seeds": "default",
    },
    "sensitivities": [
        {
            "name": "rms_seed",
            "type": "seed",
        }
    ],
    "defaults": {},
}


def test_dependencies_column_is_omitted_when_unused(tmp_path):
    output_path = tmp_path / "without-dependencies.xlsx"
    render_workbook(MINIMAL_SPEC, output_path)

    worksheet = openpyxl.load_workbook(output_path)["designinput"]
    headers = [cell.value for cell in worksheet[1]]

    assert len(headers) == 16
    assert "dependencies" not in headers


def test_render_workbook_preserves_formula_and_url_strings(tmp_path):
    output_path = tmp_path / "strings.xlsx"
    auxiliary_path = tmp_path / "auxiliary.xlsx"
    render_workbook(
        {
            **MINIMAL_SPEC,
            "defaults": {
                "FORMULA": "=1+1",
                "URL": "https://example.com",
            },
            "auxiliary_files": {
                auxiliary_path.name: {
                    "rows": [["=2+2", "https://example.org"]],
                }
            },
        },
        output_path,
    )

    defaults = openpyxl.load_workbook(output_path)["defaultvalues"]
    assert defaults["B2"].value == "=1+1"
    assert defaults["B2"].data_type == "s"
    assert defaults["B3"].value == "https://example.com"
    assert defaults["B3"].hyperlink is None

    auxiliary = openpyxl.load_workbook(auxiliary_path).active
    assert auxiliary["A1"].value == "=2+2"
    assert auxiliary["A1"].data_type == "s"
    assert auxiliary["B1"].value == "https://example.org"
    assert auxiliary["B1"].hyperlink is None


@pytest.mark.parametrize(
    ("spec", "match"),
    [
        (
            {
                **MINIMAL_SPEC,
                "auxiliary_files": {
                    "../outside.xlsx": {
                        "rows": [[1]],
                    }
                },
            },
            "simple filenames",
        ),
        (
            {
                **MINIMAL_SPEC,
                "auxiliary_files": {
                    "invalid.xlsx": {
                        "rows": [[1]],
                    }
                },
            },
            "cannot replace the main workbook",
        ),
    ],
)
def test_render_workbook_rejects_invalid_specs(tmp_path, spec, match):
    output_path = tmp_path / "invalid.xlsx"

    with pytest.raises(ValueError, match=match):
        render_workbook(spec, output_path)

    assert not output_path.exists()
